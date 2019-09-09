from openpyxl import load_workbook

workbook = load_workbook('新建 XLSX 工作表.xlsx')
sheet1 = workbook.active
cell_1_1 = sheet1.cell(1, 1)
print(cell_1_1.value)