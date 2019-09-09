from openpyxl import  Workbook
from calendar import Calendar

workbook = Workbook()

worksheet = workbook.active
workbook.remove(worksheet)

cal = Calendar()
year = 2020
for m in range(1, 13):
    title = str(m) + "月份"
    sheet = workbook.create_sheet(title)
    
    month = cal.itermonthdays2(2020, m)
    row = 1
    for day in month:
        print(day)
        if day[0] == 0:
            continue
        sheet.cell(row, 1, day[0])
        sheet.cell(row, 2, day[1])
        row += 1

workbook.save('2012年日历.xlsx')
